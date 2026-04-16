"""End-to-end XLSX export — opens the generated file and validates contents."""
import secrets
import zipfile
from pathlib import Path

from openpyxl import load_workbook

from apps.catalog.models import DatasetRow


def _approved_ds(client, code):
    ds = client.post("/api/v1/datasets",
                     {"code": code, "display_name": code}, format="json").json()
    client.post(f"/api/v1/datasets/{ds['id']}/fields",
                {"field_key": "name", "display_name": "N", "data_type": "string"},
                format="json")
    client.patch(f"/api/v1/datasets/{ds['id']}",
                 {"approval_state": "approved"}, format="json", HTTP_IF_MATCH='"1"')
    return ds


def test_xlsx_export_endpoint_produces_real_xlsx(authed_client, settings, tmp_path):
    client, _, _ = authed_client(roles=("operations",))
    ds = _approved_ds(client, "xlsx_e2e")
    DatasetRow.objects.create(dataset_id=ds["id"], payload={"name": "alpha"})
    DatasetRow.objects.create(dataset_id=ds["id"], payload={"name": "beta"})
    rdef = client.post(
        "/api/v1/reports/definitions",
        {"name": "xlsx_e2e_def", "dataset_id": ds["id"]}, format="json",
    ).json()
    settings.EXPORT_STORAGE_DIR = tmp_path / "exports"
    run = client.post(
        "/api/v1/reports/runs",
        {"report_definition_id": rdef["id"]}, format="json",
        HTTP_IDEMPOTENCY_KEY=secrets.token_hex(8),
    ).json()
    res = client.post(
        f"/api/v1/reports/runs/{run['id']}/exports",
        {"format": "xlsx"}, format="json",
        HTTP_IDEMPOTENCY_KEY=secrets.token_hex(8),
    )
    assert res.status_code == 201, res.content
    body = res.json()
    assert body["format"] == "xlsx"
    assert body["file_count"] == 1

    files = client.get(f"/api/v1/exports/{body['export_job_id']}/files").json()["files"]
    # API no longer exposes internal path — use the download endpoint instead.
    assert "path" not in files[0], "internal path must not leak in API response"
    part = files[0]["part_number"]
    dl = client.get(f"/api/v1/exports/{body['export_job_id']}/files/{part}/download")
    assert dl.status_code == 200
    # Write streamed content to a temp file so we can inspect it.
    dl_path = tmp_path / "downloaded.xlsx"
    dl_path.write_bytes(b"".join(dl.streaming_content))
    assert zipfile.is_zipfile(dl_path), "must be a real XLSX (ZIP) artifact"
    wb = load_workbook(dl_path)
    ws = wb.active
    header = [c.value for c in ws[1]]
    assert header == ["name"]
    rows = [c.value for c in ws["A"][1:]]
    assert sorted(rows) == ["alpha", "beta"]
