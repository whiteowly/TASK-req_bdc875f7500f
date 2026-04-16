"""Validate the XLSX writer produces a real Office Open XML workbook.

Loads the produced file back with openpyxl and asserts on actual cell values,
not just file existence.
"""
import zipfile
from pathlib import Path

import pytest
from openpyxl import load_workbook

pytestmark = pytest.mark.no_db

from apps.exports.services import _write_xlsx


def test_xlsx_is_valid_zip_with_required_parts(tmp_path):
    rows = [{"a": 1, "b": "x"}, {"a": 2, "b": "y"}]
    path = tmp_path / "out.xlsx"
    _write_xlsx(path, rows)
    assert path.exists()
    assert zipfile.is_zipfile(path)
    with zipfile.ZipFile(path) as z:
        names = set(z.namelist())
    assert "xl/workbook.xml" in names
    assert "xl/worksheets/sheet1.xml" in names
    assert "[Content_Types].xml" in names


def test_xlsx_contents_roundtrip(tmp_path):
    rows = [
        {"program": "CS", "gpa": 3.7},
        {"program": "Math", "gpa": 3.5},
        {"program": "Bio", "gpa": 3.1},
    ]
    path = tmp_path / "data.xlsx"
    _write_xlsx(path, rows)
    wb = load_workbook(path)
    ws = wb.active
    header = [c.value for c in ws[1]]
    assert header == ["gpa", "program"]
    body = [tuple(c.value for c in row) for row in ws.iter_rows(min_row=2)]
    assert body == [(3.7, "CS"), (3.5, "Math"), (3.1, "Bio")]


def test_xlsx_empty_rows_still_valid(tmp_path):
    path = tmp_path / "empty.xlsx"
    _write_xlsx(path, [])
    assert zipfile.is_zipfile(path)
    wb = load_workbook(path)
    assert "data" in wb.sheetnames
